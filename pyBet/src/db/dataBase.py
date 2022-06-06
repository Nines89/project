import openpyxl
from src.scrape.retakeData import PageResult

class Excel:
    def __init__(self, name):
        name = './data/' + name +'.xlsx'
        self.name = name
        self.wb = openpyxl.load_workbook(self.name)
        self.sheet = None
        self.max_row = None
        self.max_column = None

    def openSheet(self, league):
        self.sheet = self.wb[league]
        self.max_row = self.sheet.max_row
        self.max_column = self.sheet.max_column

    def printRow(self, row):
        for index in range(1, self.max_column):
            print(self.sheet.cell(row=row, column=index).value)

    def writeRow(self, mat):
        row = self.max_row + 1
        # statistiche
        self.sheet.cell(row=row, column=1).value = mat.data
        self.sheet.cell(row=row, column=2).value = mat.team1
        self.sheet.cell(row=row, column=3).value = mat.team2
        self.sheet.cell(row=row, column=4).value = mat.classifica1
        self.sheet.cell(row=row, column=5).value = mat.classifica2
        self.sheet.cell(row=row, column=6).value = mat.forma1
        self.sheet.cell(row=row, column=7).value = mat.forma2
        self.sheet.cell(row=row, column=8).value = mat.goalPT1
        self.sheet.cell(row=row, column=9).value = mat.goalPT2
        self.sheet.cell(row=row, column=10).value = mat.goalST1
        self.sheet.cell(row=row, column=11).value = mat.goalST2
        self.sheet.cell(row=row, column=12).value = mat.totGoal1
        self.sheet.cell(row=row, column=13).value = mat.totGoal2
        self.sheet.cell(row=row, column=14).value = mat.totGoal
        self.sheet.cell(row=row, column=15).value = mat.possPalla1
        self.sheet.cell(row=row, column=16).value = mat.possPalla2
        self.sheet.cell(row=row, column=17).value = mat.tiri1
        self.sheet.cell(row=row, column=18).value = mat.tiri2
        self.sheet.cell(row=row, column=19).value = mat.tiriPorta1
        self.sheet.cell(row=row, column=20).value = mat.tiriPorta2
        self.sheet.cell(row=row, column=21).value = mat.tiriFuori1
        self.sheet.cell(row=row, column=22).value = mat.tiriFuori2
        self.sheet.cell(row=row, column=23).value = mat.tiriFermati1
        self.sheet.cell(row=row, column=24).value = mat.tiriFermati2
        self.sheet.cell(row=row, column=25).value = mat.Punizioni1
        self.sheet.cell(row=row, column=26).value = mat.Punizioni2
        self.sheet.cell(row=row, column=27).value = mat.Angolo1
        self.sheet.cell(row=row, column=28).value = mat.Angolo2
        self.sheet.cell(row=row, column=29).value = mat.FuoriGioco1
        self.sheet.cell(row=row, column=30).value = mat.FuoriGioco2
        self.sheet.cell(row=row, column=31).value = mat.Parate1
        self.sheet.cell(row=row, column=32).value = mat.Parate2
        self.sheet.cell(row=row, column=33).value = mat.Falli1
        self.sheet.cell(row=row, column=34).value = mat.Falli2
        self.sheet.cell(row=row, column=35).value = mat.Rossi1
        self.sheet.cell(row=row, column=36).value = mat.Rossi2
        self.sheet.cell(row=row, column=37).value = mat.Gialli1
        self.sheet.cell(row=row, column=38).value = mat.Gialli2
        self.sheet.cell(row=row, column=39).value = mat.Passaggi1
        self.sheet.cell(row=row, column=40).value = mat.Passaggi2
        self.sheet.cell(row=row, column=41).value = mat.Contrasti1
        self.sheet.cell(row=row, column=42).value = mat.Contrasti2
        self.sheet.cell(row=row, column=43).value = mat.Attacchi1
        self.sheet.cell(row=row, column=44).value = mat.Attacchi2
        self.sheet.cell(row=row, column=45).value = mat.AttacchiPer1
        self.sheet.cell(row=row, column=46).value = mat.AttacchiPer2
        self.sheet.cell(row=row, column=47).value = mat.NumInfortunati1
        self.sheet.cell(row=row, column=48).value = mat.NumInfortunati2
        # quote 1x2
        self.sheet.cell(row=row, column=49).value = mat.quota1
        self.sheet.cell(row=row, column=50).value = mat.quotaX
        self.sheet.cell(row=row, column=51).value = mat.quota2

        # precedenti und/over
        self.sheet.cell(row=row, column=52).value = mat.numOver05_1
        self.sheet.cell(row=row, column=53).value = mat.numOver05_2
        self.sheet.cell(row=row, column=54).value = mat.numUnder05_1
        self.sheet.cell(row=row, column=55).value = mat.numUnder05_2
        self.sheet.cell(row=row, column=56).value = mat.numOver15_1
        self.sheet.cell(row=row, column=57).value = mat.numOver15_2
        self.sheet.cell(row=row, column=58).value = mat.numUnder15_1
        self.sheet.cell(row=row, column=59).value = mat.numUnder15_2
        self.sheet.cell(row=row, column=60).value = mat.numOver25_1
        self.sheet.cell(row=row, column=61).value = mat.numOver25_2
        self.sheet.cell(row=row, column=62).value = mat.numUnder25_1
        self.sheet.cell(row=row, column=63).value = mat.numUnder25_2
        self.sheet.cell(row=row, column=64).value = mat.numOver35_1
        self.sheet.cell(row=row, column=65).value = mat.numOver35_2
        self.sheet.cell(row=row, column=66).value = mat.numUnder35_1
        self.sheet.cell(row=row, column=67).value = mat.numUnder35_2
        self.sheet.cell(row=row, column=68).value = mat.numOver45_1
        self.sheet.cell(row=row, column=69).value = mat.numOver45_2
        self.sheet.cell(row=row, column=70).value = mat.numUnder45_1
        self.sheet.cell(row=row, column=71).value = mat.numUnder45_2

        # quote U/O
        self.sheet.cell(row=row, column=72).value = mat.quotaUnd05
        self.sheet.cell(row=row, column=73).value = mat.quotaOv05
        self.sheet.cell(row=row, column=74).value = mat.quotaUnd15
        self.sheet.cell(row=row, column=75).value = mat.quotaOv15
        self.sheet.cell(row=row, column=76).value = mat.quotaUnd25
        self.sheet.cell(row=row, column=77).value = mat.quotaOv25
        self.sheet.cell(row=row, column=78).value = mat.quotaUnd35
        self.sheet.cell(row=row, column=79).value = mat.quotaOv35
        self.sheet.cell(row=row, column=80).value = mat.quotaUnd45
        self.sheet.cell(row=row, column=81).value = mat.quotaOv45

        # quote doppia chance
        self.sheet.cell(row=row, column=82).value = mat.quota12
        self.sheet.cell(row=row, column=83).value = mat.quota1X
        self.sheet.cell(row=row, column=84).value = mat.quotaX2

        #quote GG/NG
        self.sheet.cell(row=row, column=85).value = mat.quotaGG
        self.sheet.cell(row=row, column=86).value = mat.quotaNG

        #last match
        self.sheet.cell(row=row, column=87).value = str(mat.lastMatchs1)
        self.sheet.cell(row=row, column=88).value = str(mat.lastMatchs2)
        self.sheet.cell(row=row, column=89).value = str(mat.prevMatchsVS)
        self.sheet.cell(row=row, column=90).value = str(mat.lastMatchsCASA)
        self.sheet.cell(row=row, column=91).value = str(mat.lastMatchsTRASFERTA)

    def saveCVS(self):
        self.wb.save(self.name)

    def readRow(self, row):
        mat = PageResult()
        # statistiche
        mat.data =self.sheet.cell(row=row, column=1).value
        mat.team1 = self.sheet.cell(row=row, column=2).value
        mat.team2 = self.sheet.cell(row=row, column=3).value
        mat.classifica1 = self.sheet.cell(row=row, column=4).value
        mat.classifica2 = self.sheet.cell(row=row, column=5).value
        mat.forma1 = self.sheet.cell(row=row, column=6).value
        mat.forma2 = self.sheet.cell(row=row, column=7).value
        mat.goalPT1 = self.sheet.cell(row=row, column=8).value
        mat.goalPT2 = self.sheet.cell(row=row, column=9).value
        mat.goalST1 = self.sheet.cell(row=row, column=10).value
        mat.goalST2 = self.sheet.cell(row=row, column=11).value
        mat.totGoal1 = self.sheet.cell(row=row, column=12).value
        mat.totGoal2 = self.sheet.cell(row=row, column=13).value
        mat.totGoal = self.sheet.cell(row=row, column=14).value
        mat.possPalla1 = self.sheet.cell(row=row, column=15).value
        mat.possPalla2 = self.sheet.cell(row=row, column=16).value
        mat.tiri1 = self.sheet.cell(row=row, column=17).value
        mat.tiri2 = self.sheet.cell(row=row, column=18).value
        mat.tiriPorta1 = self.sheet.cell(row=row, column=19).value
        mat.tiriPorta2 = self.sheet.cell(row=row, column=20).value
        mat.tiriFuori1 = self.sheet.cell(row=row, column=21).value
        mat.tiriFuori2 = self.sheet.cell(row=row, column=22).value
        mat.tiriFermati1 = self.sheet.cell(row=row, column=23).value
        mat.tiriFermati2 = self.sheet.cell(row=row, column=24).value
        mat.Punizioni1 = self.sheet.cell(row=row, column=25).value
        mat.Punizioni2 = self.sheet.cell(row=row, column=26).value
        mat.Angolo1 = self.sheet.cell(row=row, column=27).value
        mat.Angolo2 = self.sheet.cell(row=row, column=28).value
        mat.FuoriGioco1 = self.sheet.cell(row=row, column=29).value
        mat.FuoriGioco2 = self.sheet.cell(row=row, column=30).value
        mat.Parate1 = self.sheet.cell(row=row, column=31).value
        mat.Parate2 = self.sheet.cell(row=row, column=32).value
        mat.Falli1 = self.sheet.cell(row=row, column=33).value
        mat.Falli2 = self.sheet.cell(row=row, column=34).value
        mat.Rossi1 = self.sheet.cell(row=row, column=35).value
        mat.Rossi2 = self.sheet.cell(row=row, column=36).value
        mat.Gialli1 = self.sheet.cell(row=row, column=37).value
        mat.Gialli2 = self.sheet.cell(row=row, column=38).value
        mat.Passaggi1 = self.sheet.cell(row=row, column=39).value
        mat.Passaggi2 = self.sheet.cell(row=row, column=40).value
        mat.Contrasti1 = self.sheet.cell(row=row, column=41).value
        mat.Contrasti2 = self.sheet.cell(row=row, column=42).value
        mat.Attacchi1 = self.sheet.cell(row=row, column=43).value
        mat.Attacchi2 = self.sheet.cell(row=row, column=44).value
        mat.AttacchiPer1 = self.sheet.cell(row=row, column=45).value
        mat.AttacchiPer2 = self.sheet.cell(row=row, column=46).value
        mat.NumInfortunati1 = self.sheet.cell(row=row, column=47).value
        mat.NumInfortunati2 = self.sheet.cell(row=row, column=48).value
        # quote 1x2
        mat.quota1 = self.sheet.cell(row=row, column=49).value
        mat.quotaX = self.sheet.cell(row=row, column=50).value
        mat.quota2 = self.sheet.cell(row=row, column=51).value

        # precedenti und/over
        mat.numOver05_1 = self.sheet.cell(row=row, column=52).value
        mat.numOver05_2 = self.sheet.cell(row=row, column=53).value
        mat.numUnder05_1 = self.sheet.cell(row=row, column=54).value
        mat.numUnder05_2 = self.sheet.cell(row=row, column=55).value
        mat.numOver15_1 = self.sheet.cell(row=row, column=56).value
        mat.numOver15_2 = self.sheet.cell(row=row, column=57).value
        mat.numUnder15_1 = self.sheet.cell(row=row, column=58).value
        mat.numUnder15_2 = self.sheet.cell(row=row, column=59).value
        mat.numOver25_1 = self.sheet.cell(row=row, column=60).value
        mat.numOver25_2 = self.sheet.cell(row=row, column=61).value
        mat.numUnder25_1 = self.sheet.cell(row=row, column=62).value
        mat.numUnder25_2 = self.sheet.cell(row=row, column=63).value
        mat.numOver35_1 = self.sheet.cell(row=row, column=64).value
        mat.numOver35_2 = self.sheet.cell(row=row, column=65).value
        mat.numUnder35_1 = self.sheet.cell(row=row, column=66).value
        mat.numUnder35_2 = self.sheet.cell(row=row, column=67).value
        mat.numOver45_1 = self.sheet.cell(row=row, column=68).value
        mat.numOver45_2 = self.sheet.cell(row=row, column=69).value
        mat.numUnder45_1 = self.sheet.cell(row=row, column=70).value
        mat.numUnder45_2 = self.sheet.cell(row=row, column=71).value

        # quote U/O
        mat.quotaUnd05 = self.sheet.cell(row=row, column=72).value
        mat.quotaOv05 = self.sheet.cell(row=row, column=73).value
        mat.quotaUnd15 = self.sheet.cell(row=row, column=74).value
        mat.quotaOv15 = self.sheet.cell(row=row, column=75).value
        mat.quotaUnd25 = self.sheet.cell(row=row, column=76).value
        mat.quotaOv25 = self.sheet.cell(row=row, column=77).value
        mat.quotaUnd35 = self.sheet.cell(row=row, column=78).value
        mat.quotaOv35 = self.sheet.cell(row=row, column=79).value
        mat.quotaUnd45 = self.sheet.cell(row=row, column=80).value
        mat.quotaOv45 = self.sheet.cell(row=row, column=81).value

        # quote doppia chance
        mat.quota12 = self.sheet.cell(row=row, column=82).value
        mat.quota1X = self.sheet.cell(row=row, column=83).value
        mat.quotaX2 = self.sheet.cell(row=row, column=84).value

        #quote GG/NG
        mat.quotaGG = self.sheet.cell(row=row, column=85).value
        mat.quotaNG = self.sheet.cell(row=row, column=86).value

        #last match
        mat.lastMatchs1 = self.sheet.cell(row=row, column=87).value
        mat.lastMatchs2 = self.sheet.cell(row=row, column=88).value
        mat.prevMatchsVS = self.sheet.cell(row=row, column=89).value
        mat.lastMatchsCASA = self.sheet.cell(row=row, column=90).value
        mat.lastMatchsTRASFERTA = self.sheet.cell(row=row, column=91).value
        return mat

    def readAllRows(self):
        """
        Def of a function = def readAllRows(self) -> List[PageResult]
        Legge tutte le partite di un singolo excel

        :return: Lista
        Ogni elemento della lista rappresenta una partita
        """
        matches = []
        for sheet in self.wb.worksheets:
            for i in range(2, sheet.max_row):
                matches.append(self.readRow(i))
        return matches

if __name__ == "__main__":
    f = Excel('data')
    f.openSheet('Serie A') # apre il foglio
    mat = f.readAllRows()
    print(mat[0].team1)
    print(len(mat))

    #SerieA.cell(row=SerieA.max_row + 1,
    #            column=SerieA.max_column).value = "Ciao"  # inserisce nell'ultima colonna della riga successiva
    # wb.save('data.xlsx') #salva il file
